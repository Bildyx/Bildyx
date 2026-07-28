#!/usr/bin/env python3
"""
Generates one Excel import template per targeted Prisma model.

Reads apps/api/prisma/schema.prisma, keeps only the "reference data" models
(industries, countries, cities, jobs, organizations, skills, certifications,
study fields, degrees) and writes an .xlsx file per model with the model's
attributes as column headers, ready to be filled in and imported into the
database. Relation fields (back-references to other models) and
auto-managed columns (id, created_at, updated_at, deleted_at) are skipped
since they aren't meant to be filled in manually - except implicit
many-to-many relations (e.g. Organization<->Country), which have no
foreign-key column anywhere in the schema to carry them. Those get a single
free-text ";"-separated column instead (see M2M_COLUMNS), resolved by
name/key at import time the same way other array columns already are.

Scalar foreign keys (e.g. Organization.city_id, a raw UUID column in the DB)
are NEVER shown to the client as-is: the template exposes the *target
model's natural key* instead (e.g. "city_serial_number"), resolved via
FK_TARGET_MODEL below. This is what makes the whole pipeline usable by
someone who shouldn't ever have to type or look up a UUID.

Every generated template is locked down against import errors: the header
row and any column outside the defined fields are protected (no limit on
the number of rows - only the set of columns is restricted), row/column
insertion and deletion is blocked, and each column gets a data validation
matching its Prisma type (whole number, decimal, date, or enum dropdown)
with an English error message. Sheet protection and workbook structure
locking are both password-protected with TEMPLATE_PASSWORD.

Two extra modes support downloading a template PRE-FILLED with the current
database content (rather than an empty one):

  --dump-spec PATH
      Instead of writing .xlsx files, writes a JSON description of every
      target model's columns (Prisma field name, CSV/template column name,
      and how to resolve it - plain scalar, list, FK, or M2M) to PATH. This
      JSON is the only thing scripts/export-current-data.ts needs to know
      to query Prisma and produce a matching CSV per model, with no
      per-model logic duplicated between Python and TypeScript.

  --populate-from-db DIR
      Looks for a "<table_name>.csv" file per model in DIR (the output of
      export-current-data.ts) and writes it into the generated workbook
      starting at row 2, with the exact same locking/validation as an empty
      template. Implies --force (the whole point is to refresh the
      downloadable file with current data).

Usage:
    python3 generate_excel_templates.py [--schema PATH] [--output DIR]
    python3 generate_excel_templates.py --protect-existing [--schema PATH] [--output DIR]
    python3 generate_excel_templates.py --dump-spec ./template-spec.json
    python3 generate_excel_templates.py --populate-from-db ./data/exports
"""

import argparse
import csv as csv_module
import json
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.worksheet.worksheet import Worksheet

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_SCHEMA = SCRIPT_DIR.parent / "prisma" / "schema.prisma"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "data" / "excel_templates"

TARGET_MODELS = [
    "Industry",
    "Country",
    "City",
    "Job",
    "Organization",
    "Skill",
    "Certification",
    # "University" retiré : le modèle a été supprimé du schéma, ses champs
    # (studentCount, undergraduates, postgraduates) vivent directement sur
    # Organization désormais - voir la conversation sur la fusion
    # University -> Organization.
    "StudyFields",
    "Degree",
    "Subject",
]

# Auto-managed columns: never filled in manually.
EXCLUDED_FIELDS = {
    "id",
    "createdAt",
    "created_at",
    "updatedAt",
    "updated_at",
    "deletedAt",
    "deleted_at",
}

# Implicit many-to-many relations (Prisma list fields with no `fields:`/
# `references:` on either side, backed by a hidden join table) have no
# foreign-key column to expose. Each such relation is exposed as a single
# free-text column on exactly one of its two models (whichever is the more
# natural place to edit it from), keyed by (model_name, schema_field_name).
# The Industry self-relation only needs one of its two fields (industries_A/
# industries_B) since both describe the same symmetric relation.
M2M_COLUMNS = {
    ("Organization", "countries"): "countries",
    ("Organization", "industries"): "industries",
    ("Organization", "cities_working_area"): "working_area_cities",
    ("City", "mainIndustries"): "main_industries",
    ("Subject", "industries"): "industries",
    ("Industry", "industries_A"): "related_industries",
}

# Every target model's natural (human-meaningful, stable, client-facing)
# key field. Used two ways: (1) to rename a scalar FK column so it shows the
# target's natural key instead of a raw UUID - irrelevant for Country since
# none of its scalar FK columns are @db.Uuid in the first place (its @id IS
# already the natural key, isoCode, so e.g. City.countryId is @db.Char(2)
# and never flagged as is_uuid_fk); and (2) to know which field to pull for
# M2M/list relations pointing at that model (e.g. Organization.countries),
# where Country's entry below IS needed.
NATURAL_KEY_FIELD_BY_MODEL = {
    "Organization": "slug",
    "City": "serial_number",
    "Industry": "serial_number",
    "Country": "isoCode",
}

# Prisma field name (as used in the schema, not the DB column) holding the
# soft-delete timestamp for each target model. Used only by --dump-spec, so
# export-current-data.ts can filter out soft-deleted rows without needing to
# know each model's naming quirk itself (mirrors ImportAdapter.deletedAtField
# in prisma/import/types.ts).
DELETED_AT_FIELD_BY_MODEL = {
    "Industry": "deletedAt",
    "Country": "deletedAt",
    "City": "deletedAt",
    "Job": "deletedAt",
    "Organization": "deletedAt",
    "Skill": "deletedAt",
    "Certification": "deletedAt",
    "StudyFields": "deleted_at",
    "Degree": "deletedAt",
    "Subject": "deleted_at",
}

# Excel's actual row limit. Data validation ranges are cheap to extend this
# far (stored as a single range reference, not per-cell), so every data
# column is validated/unlockable for its entire length - only the column
# (width) is restricted, never the number of rows.
EXCEL_MAX_ROW = 1_048_576

LISTS_SHEET_NAME = "_lists"

# Password protecting sheet protection and workbook structure locking below,
# so unprotecting a template requires knowing it rather than a single click.
TEMPLATE_PASSWORD = "bildyx"

MODEL_RE = re.compile(r"model\s+(\w+)\s*\{(.*?)\n\}", re.S)
ENUM_RE = re.compile(r"enum\s+(\w+)\s*\{(.*?)\n\}", re.S)
FIELD_LINE_RE = re.compile(r"^([A-Za-z_]\w*)\s+([A-Za-z_]\w*(?:\[\])?\??)\s*(.*)$")
MAP_ATTR_RE = re.compile(r'@map\("([^"]+)"\)')
TABLE_MAP_RE = re.compile(r'@@map\("([^"]+)"\)')
RELATION_FIELDS_RE = re.compile(r"@relation\([^)]*fields:\s*\[([^\]]+)\][^)]*\)")


def camel_to_snake(name: str) -> str:
    s = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", name)
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", s)
    return s.lower()


def prisma_client_accessor(model_name: str) -> str:
    """PrismaClient delegate name for a model: first letter lowercased,
    rest unchanged (e.g. "StudyFields" -> "studyFields")."""
    return model_name[0].lower() + model_name[1:]


def parse_blocks(schema_text: str, pattern: re.Pattern) -> dict[str, str]:
    return {m.group(1): m.group(2) for m in pattern.finditer(schema_text)}


def parse_enum_values(body: str) -> list[str]:
    values = []
    for line in body.splitlines():
        line = line.strip()
        if not line or line.startswith("//"):
            continue
        values.append(line.split()[0])
    return values


def parse_field_type(raw_type: str) -> tuple[str, bool, bool]:
    """Returns (base_type, is_optional, is_list)."""
    is_optional = raw_type.endswith("?")
    if is_optional:
        raw_type = raw_type[:-1]
    is_list = raw_type.endswith("[]")
    if is_list:
        raw_type = raw_type[:-2]
    return raw_type, is_optional, is_list


def find_fk_target_model(body: str, fk_field_name: str) -> str | None:
    """
    Given a scalar FK field's own Prisma name (e.g. "parentOrganizationId"),
    finds the model it points to by locating the relation line that
    references it (`xxx Model? @relation(fields: [parentOrganizationId], ...)`)
    - fields: always lists the scalar's Prisma field name, never its mapped
    DB column name, so this must be matched against field_name, not
    column_name.
    """
    for line in body.splitlines():
        line = line.strip()
        if not line or line.startswith("//") or line.startswith("@@"):
            continue
        rel_match = RELATION_FIELDS_RE.search(line)
        if not rel_match:
            continue
        referenced = [f.strip() for f in rel_match.group(1).split(",")]
        if fk_field_name in referenced:
            m = FIELD_LINE_RE.match(line)
            if m:
                return m.group(2).rstrip("?")
    return None


def resolve_fk_column(
    column_name: str, target_model: str | None, target_natural_key_field: str | None
) -> str:
    """
    Renamed column shown in the template/export for a scalar FK, pointing at
    the target's natural key rather than the raw UUID
    (e.g. "city_id" -> "city_serial_number"). Falls back to the original
    column name if the target model isn't in NATURAL_KEY_FIELD_BY_MODEL
    (Country's FKs, which already hold a natural key, and any model not yet
    added to the map - printed as a warning by the caller in that case).
    """
    if not target_model or not target_natural_key_field:
        return column_name
    base = column_name
    if base.endswith("_id"):
        base = base[: -len("_id")]
    elif base.endswith("Id"):
        base = camel_to_snake(base[: -len("Id")])
    suffix = camel_to_snake(target_natural_key_field)
    return f"{base}_{suffix}"


def parse_model_fields(
    body: str, model_names: set[str], model_name: str
) -> list[dict]:
    fields = []
    for line in body.splitlines():
        line = line.strip()
        if not line or line.startswith("//") or line.startswith("@@"):
            continue
        m = FIELD_LINE_RE.match(line)
        if not m:
            continue
        field_name, raw_type, attrs = m.groups()
        base_type, is_optional, is_list = parse_field_type(raw_type)

        # Fields whose type is another model are relations, not real columns
        # - except the curated implicit many-to-many relations in
        # M2M_COLUMNS, which get a free-text ";"-separated column since they
        # have no foreign-key column anywhere to represent them instead.
        if base_type in model_names:
            m2m_column = M2M_COLUMNS.get((model_name, field_name))
            if m2m_column is None:
                continue
            target_natural_key_field = NATURAL_KEY_FIELD_BY_MODEL.get(base_type)
            fields.append(
                {
                    "column": m2m_column,
                    "field_name": field_name,
                    "base_type": base_type,
                    "is_list": is_list,
                    "required": False,
                    "is_uuid_fk": False,
                    "is_m2m": True,
                    "target_model": base_type,
                    "target_natural_key_field": target_natural_key_field,
                }
            )
            continue

        map_match = MAP_ATTR_RE.search(attrs)
        column_name = map_match.group(1) if map_match else field_name
        if field_name in EXCLUDED_FIELDS or column_name in EXCLUDED_FIELDS:
            continue

        required = not is_optional and not is_list and "@default(" not in attrs
        is_uuid_fk = "@db.Uuid" in attrs and (
            field_name.endswith("Id") or field_name.endswith("_id")
        )

        target_model = None
        target_natural_key_field = None
        resolved_column = column_name
        if is_uuid_fk:
            target_model = find_fk_target_model(body, field_name)
            target_natural_key_field = NATURAL_KEY_FIELD_BY_MODEL.get(target_model or "")
            if target_model and target_model not in NATURAL_KEY_FIELD_BY_MODEL and target_model != "Country":
                print(
                    f"  Warning: {model_name}.{field_name} -> {target_model}, but "
                    f"{target_model} has no entry in NATURAL_KEY_FIELD_BY_MODEL. "
                    f"Column left as raw id ('{column_name}') - add it if this FK "
                    f"should be human-fillable."
                )
            resolved_column = resolve_fk_column(
                column_name, target_model, target_natural_key_field
            )

        fields.append(
            {
                "column": resolved_column,
                "field_name": field_name,
                "base_type": base_type,
                "is_list": is_list,
                "required": required,
                "is_uuid_fk": is_uuid_fk,
                "is_m2m": False,
                "target_model": target_model,
                "target_natural_key_field": target_natural_key_field,
            }
        )
    return fields


def number_format_for(field: dict, enums: dict[str, list[str]]) -> str | None:
    """Excel number_format matching the field's Prisma type, or None for General."""
    base_type = field["base_type"]

    if field["is_list"] or field.get("is_uuid_fk") or field.get("is_m2m") or base_type in enums:
        return "@"
    if base_type in ("Int", "BigInt"):
        return "#,##0"
    if base_type in ("Float", "Decimal"):
        return "#,##0.00"
    if base_type == "DateTime":
        return "yyyy-mm-dd"
    if base_type == "Boolean":
        return None
    return "@"


def table_name_for(model_name: str, body: str) -> str:
    map_match = TABLE_MAP_RE.search(body)
    return map_match.group(1) if map_match else camel_to_snake(model_name)


def write_lists_sheet(wb: Workbook, enums_used: dict[str, list[str]]) -> dict[str, str]:
    """
    Writes each enum's values into a hidden helper sheet and returns a map of
    enum name -> range reference, so list validations can point at a range
    instead of an inline comma string. Inline lists are capped at 255
    characters by Excel, which silently drops the dropdown for large enums
    (e.g. Currency); a range reference has no such limit.
    """
    if not enums_used:
        return {}

    lists_ws = wb.create_sheet(LISTS_SHEET_NAME)
    refs = {}
    for col_idx, (enum_name, values) in enumerate(sorted(enums_used.items()), start=1):
        col_letter = lists_ws.cell(row=1, column=col_idx).column_letter
        lists_ws.cell(row=1, column=col_idx, value=enum_name)
        for row_idx, value in enumerate(values, start=2):
            lists_ws.cell(row=row_idx, column=col_idx, value=value)
        last_row = len(values) + 1
        refs[enum_name] = f"{LISTS_SHEET_NAME}!${col_letter}$2:${col_letter}${last_row}"

    lists_ws.sheet_state = "veryHidden"
    return refs


def data_validation_for(field: dict, enums: dict, list_refs: dict[str, str]) -> DataValidation | None:
    """
    Data validation enforcing only the field's Prisma type, with an English
    error message. Blank is always allowed, even for required fields - the
    template shouldn't block partially-filled rows, only wrong-typed values.
    FK/M2M columns (now holding natural-key text, not the original scalar
    type) never get numeric/date validation even if the underlying UUID
    column's Prisma type happened to look numeric-adjacent.
    """
    if field.get("is_uuid_fk") or field.get("is_m2m"):
        return None

    base_type = field["base_type"]

    if base_type in enums and not field["is_list"]:
        return DataValidation(
            type="list",
            formula1=list_refs[base_type],
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error="Please select a value from the dropdown list.",
        )
    if field["is_list"]:
        return None
    if base_type in ("Int", "BigInt"):
        return DataValidation(
            type="whole",
            operator="between",
            formula1=-2147483648,
            formula2=2147483647,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid number",
            error="Please enter a whole number.",
        )
    if base_type in ("Float", "Decimal"):
        return DataValidation(
            type="decimal",
            operator="between",
            formula1=-999999999999.0,
            formula2=999999999999.0,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid number",
            error="Please enter a numeric value.",
        )
    if base_type == "DateTime":
        return DataValidation(
            type="date",
            operator="between",
            formula1=date(1900, 1, 1),
            formula2=date(2999, 12, 31),
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid date",
            error="Please enter a valid date (YYYY-MM-DD).",
        )
    if base_type == "Boolean":
        return DataValidation(
            type="list",
            formula1='"TRUE,FALSE"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error="Please select TRUE or FALSE.",
        )
    return None


def enable_sheet_protection(ws: Worksheet) -> None:
    """
    Turns on sheet protection with Excel's default lockdown (blocks
    inserting/deleting rows and columns, sorting, etc.), except column/row
    resizing stays allowed - purely cosmetic (doesn't let anyone write
    outside the authorized cells or touch the data validation), and users
    otherwise can't widen a column to see long text or adjust row height.
    Protected with TEMPLATE_PASSWORD so removing it isn't a single click.
    """
    ws.protection.set_password(TEMPLATE_PASSWORD)
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def lock_down_sheet(
    ws: Worksheet, fields: list[dict], enums: dict, list_refs: dict[str, str]
) -> None:
    """
    Locks every column outside the defined fields, adds a type-matching data
    validation to each data column (unlocked/validated for its full height,
    row 2 through Excel's row limit), and enables sheet protection. Cells
    are locked by default in Excel, so only the intended data-entry columns
    need to be explicitly unlocked; everything else (header row, columns
    past the last field) is protected for free once sheet protection is
    enabled. Unlocking is done at the column level (not per-cell) so it
    covers every row without creating a cell object per row.
    """
    for col_idx, field in enumerate(fields, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].protection = Protection(locked=False)

        dv = data_validation_for(field, enums, list_refs)
        if dv:
            ws.add_data_validation(dv)
            dv.add(f"{column_letter}2:{column_letter}{EXCEL_MAX_ROW}")

    enable_sheet_protection(ws)


def coerce_value(raw: str, field: dict):
    """
    Casts a raw CSV string to the type Excel/openpyxl should store the cell
    as, so numeric/date columns sort and filter correctly instead of being
    stored as text that merely looks numeric. Falls back to the raw string
    on any parse failure rather than dropping the value silently - a
    visibly-wrong cell is easier for the client to spot and fix than a
    silently-emptied one.
    """
    if raw is None or raw == "":
        return None
    if field["is_list"] or field.get("is_m2m") or field.get("is_uuid_fk"):
        return raw
    base_type = field["base_type"]
    try:
        if base_type in ("Int", "BigInt"):
            return int(raw)
        if base_type in ("Float", "Decimal"):
            return float(raw)
        if base_type == "Boolean":
            return raw.strip().upper() in ("TRUE", "1")
    except ValueError:
        return raw
    return raw


def load_data_rows(data_dir: Path, table_name: str) -> list[dict[str, str]] | None:
    csv_path = data_dir / f"{table_name}.csv"
    if not csv_path.exists():
        return None
    with csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv_module.DictReader(f, delimiter=";")
        return list(reader)


def write_workbook(
    table_name: str,
    fields: list[dict],
    enums: dict,
    output_dir: Path,
    data_rows: list[dict[str, str]] | None = None,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]

    header_font = Font(bold=True)
    required_fill = PatternFill("solid", fgColor="FFF2CC")

    enums_used = {
        field["base_type"]: enums[field["base_type"]]
        for field in fields
        if field["base_type"] in enums and not field["is_list"] and not field.get("is_uuid_fk")
    }
    list_refs = write_lists_sheet(wb, enums_used)

    for col_idx, field in enumerate(fields, start=1):
        header_cell = ws.cell(row=1, column=col_idx, value=field["column"])
        header_cell.font = header_font
        if field["required"]:
            header_cell.fill = required_fill

        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = max(18, len(field["column"]) + 4)

        number_format = number_format_for(field, enums)
        if number_format:
            ws.column_dimensions[column_letter].number_format = number_format

    if data_rows:
        for row_idx, data_row in enumerate(data_rows, start=2):
            for col_idx, field in enumerate(fields, start=1):
                raw_value = data_row.get(field["column"], "")
                ws.cell(row=row_idx, column=col_idx, value=coerce_value(raw_value, field))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(fields)).column_letter}1"

    lock_down_sheet(ws, fields, enums, list_refs)
    wb.security.lockStructure = True
    wb.security.set_workbook_password(TEMPLATE_PASSWORD)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{table_name}.xlsx"
    wb.save(output_path)
    return output_path


def protect_existing_workbook(
    path: Path, fields: list[dict], enums: dict
) -> None:
    """Adds locking/validation to an already-filled template without touching any cell value."""
    wb = load_workbook(path)
    ws = wb.active

    headers = {
        cell.value: idx
        for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)
    }

    # Newly-added M2M columns (see M2M_COLUMNS) are appended as new trailing
    # columns so existing filled rows and column order are left untouched.
    # Other schema fields missing from the header (pre-existing drift, e.g. a
    # template that hasn't been regenerated since a schema change) are left
    # alone and just reported below, same as before.
    header_font = Font(bold=True)
    next_col = ws.max_column + 1
    for field in fields:
        if not field.get("is_m2m") or field["column"] in headers:
            continue
        col_letter = ws.cell(row=1, column=next_col).column_letter
        header_cell = ws.cell(row=1, column=next_col, value=field["column"])
        header_cell.font = header_font
        ws.column_dimensions[col_letter].width = max(18, len(field["column"]) + 4)
        number_format = number_format_for(field, enums)
        if number_format:
            ws.column_dimensions[col_letter].number_format = number_format
        headers[field["column"]] = next_col
        next_col += 1

    matched_fields = [f for f in fields if f["column"] in headers]
    missing = [f["column"] for f in fields if f["column"] not in headers]
    if missing:
        print(f"  Warning: columns not found in {path.name}, skipping: {missing}")

    # Re-derive each field's actual column index from the file's own header
    # row rather than assuming it matches schema order.
    ordered_fields = sorted(matched_fields, key=lambda f: headers[f["column"]])

    # Drop any pre-existing validations so re-running this doesn't stack
    # duplicate/overlapping rules on top of the ones added below.
    ws.data_validations = DataValidationList()

    enums_used = {
        field["base_type"]: enums[field["base_type"]]
        for field in ordered_fields
        if field["base_type"] in enums and not field["is_list"] and not field.get("is_uuid_fk")
    }
    if LISTS_SHEET_NAME in wb.sheetnames:
        del wb[LISTS_SHEET_NAME]
    list_refs = write_lists_sheet(wb, enums_used)

    existing_max_row = ws.max_row

    for field in ordered_fields:
        col_idx = headers[field["column"]]
        column_letter = ws.cell(row=1, column=col_idx).column_letter

        # Column-level unlock covers every future row (no per-cell object,
        # so it reaches all 1,048,576 rows at no cost). It does NOT reach
        # rows that already have their own explicit cell style (existing
        # filled data commonly does), so those need unlocking individually
        # too - cheap since it's bounded by the sheet's current, real size.
        ws.column_dimensions[column_letter].protection = Protection(locked=False)
        for row in range(2, existing_max_row + 1):
            ws.cell(row=row, column=col_idx).protection = Protection(locked=False)

        dv = data_validation_for(field, enums, list_refs)
        if dv:
            ws.add_data_validation(dv)
            dv.add(f"{column_letter}2:{column_letter}{EXCEL_MAX_ROW}")

    enable_sheet_protection(ws)
    if wb.security is None:
        wb.security = WorkbookProtection()
    wb.security.lockStructure = True
    wb.security.set_workbook_password(TEMPLATE_PASSWORD)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--schema", type=Path, default=DEFAULT_SCHEMA)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing .xlsx files instead of skipping them",
    )
    parser.add_argument(
        "--protect-existing",
        action="store_true",
        help=(
            "Instead of generating new templates, lock down the existing "
            ".xlsx files in --output in place (data is preserved, only "
            "protection/validation is added or refreshed)"
        ),
    )
    parser.add_argument(
        "--dump-spec",
        type=Path,
        help=(
            "Write a JSON field spec to PATH instead of generating "
            "templates - consumed by scripts/export-current-data.ts to "
            "know what to query from the DB and how to shape each column."
        ),
    )
    parser.add_argument(
        "--populate-from-db",
        type=Path,
        help=(
            "Directory containing '<table_name>.csv' exports (see "
            "scripts/export-current-data.ts) to pre-fill each generated "
            "template with the database's current content. Implies "
            "--force."
        ),
    )
    args = parser.parse_args()

    schema_text = args.schema.read_text()
    model_bodies = parse_blocks(schema_text, MODEL_RE)
    enum_bodies = parse_blocks(schema_text, ENUM_RE)
    model_names = set(model_bodies.keys())
    enums = {name: parse_enum_values(body) for name, body in enum_bodies.items()}

    spec: dict[str, dict] = {}

    for model_name in TARGET_MODELS:
        if model_name not in model_bodies:
            print(f"Skipping {model_name}: not found in schema")
            continue

        body = model_bodies[model_name]
        table_name = table_name_for(model_name, body)
        output_path = args.output / f"{table_name}.xlsx"
        fields = parse_model_fields(body, model_names, model_name)
        if not fields:
            print(f"Skipping {model_name}: no fillable fields found")
            continue

        if args.dump_spec:
            spec[model_name] = {
                "prismaModel": prisma_client_accessor(model_name),
                "tableName": table_name,
                "deletedAtField": DELETED_AT_FIELD_BY_MODEL.get(model_name, "deletedAt"),
                "fields": [
                    {
                        "column": f["column"],
                        "fieldName": f["field_name"],
                        "kind": (
                            "m2m"
                            if f.get("is_m2m")
                            else "fk"
                            if f.get("is_uuid_fk") and f.get("target_natural_key_field")
                            else "list"
                            if f["is_list"]
                            else "scalar"
                        ),
                        **(
                            {"targetNaturalKeyField": f["target_natural_key_field"]}
                            if f.get("target_natural_key_field")
                            else {}
                        ),
                    }
                    for f in fields
                ],
            }
            continue

        if args.protect_existing:
            if not output_path.exists():
                print(f"Skipping {model_name}: {output_path} does not exist")
                continue
            protect_existing_workbook(output_path, fields, enums)
            print(f"{model_name} -> {output_path} (protected)")
            continue

        data_rows = None
        if args.populate_from_db:
            data_rows = load_data_rows(args.populate_from_db, table_name)
            if data_rows is None:
                print(f"  Note: no {table_name}.csv found in {args.populate_from_db}, template will be empty")

        if output_path.exists() and not args.force and not args.populate_from_db:
            print(
                f"Skipping {model_name}: {output_path} already exists (use --force to overwrite)"
            )
            continue

        write_workbook(table_name, fields, enums, args.output, data_rows=data_rows)
        row_note = f", {len(data_rows)} lignes" if data_rows else ""
        print(f"{model_name} -> {output_path} ({len(fields)} columns{row_note})")

    if args.dump_spec:
        args.dump_spec.write_text(json.dumps(spec, indent=2), encoding="utf-8")
        print(f"Spec écrite -> {args.dump_spec} ({len(spec)} modèles)")


if __name__ == "__main__":
    main()
