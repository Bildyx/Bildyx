#!/usr/bin/env python3
"""
Converts every .xlsx file in apps/api/data/excel into a .csv file in
apps/api/data (one CSV per workbook, using its active sheet).

Usage:
    python3 excel_to_csv.py [--input DIR] [--output DIR]
"""

import argparse
import csv
import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

# Technical sheet written by generate_excel_templates.py to feed the
# dropdown lists. It must never be mistaken for the data sheet.
LISTS_SHEET_NAME = "_lists"

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR.parent / "data" / "excel_templates"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "data"


def pick_data_sheet(wb, expected_title: str):
    """
    The workbook's data sheet.

    wb.active was used until now: if the client saved the file with the
    "_lists" tab selected, the conversion overwrote the business CSV with the
    list of enum values. The anti-shrinking guard did not protect against
    that case (enums have more rows than most tables). We therefore target
    the sheet by name, falling back to the first non-technical sheet.
    """
    if expected_title in wb.sheetnames:
        return wb[expected_title]
    for name in wb.sheetnames:
        if name != LISTS_SHEET_NAME:
            return wb[name]
    raise ValueError(f"aucune feuille de données dans le classeur (onglets: {wb.sheetnames})")


def normalize_cell(value):
    """
    Renders a cell in exactly the form the import engine expects.

    openpyxl returns Excel's native types: csv.writer therefore wrote
    "1234.0" for an integer stored as a float, "2024-01-01 00:00:00" for a
    date, and "True"/"False" (Python capitalisation) for a boolean - three
    forms that toStringArray/toBool/checkInt on the import side do not
    recognise.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, dt.datetime):
        return value.date().isoformat() if value.time() == dt.time() else value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def existing_data_row_count(csv_path: Path) -> int:
    if not csv_path.exists():
        return 0
    with csv_path.open("r", newline="", encoding="utf-8") as f:
        # -1 for the header row; a file with only a header (or empty) counts as 0.
        return max(sum(1 for _ in csv.reader(f, delimiter=";")) - 1, 0)


def convert(input_dir: Path, output_dir: Path, force: bool = False) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_files = sorted(input_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"No .xlsx files found in {input_dir}")
        return

    for excel_path in excel_files:
        csv_path = output_dir / f"{excel_path.stem}.csv"
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        sheet = pick_data_sheet(wb, excel_path.stem)

        new_rows = []
        for row in sheet.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            new_rows.append([normalize_cell(cell) for cell in row])
        wb.close()

        # Empty trailing columns, inherited from a stray ";" in the
        # workbook: they are not part of the header and would fail the
        # import's strict column validation.
        if new_rows:
            width = len(new_rows[0])
            while width > 0 and new_rows[0][width - 1] == "":
                width -= 1
            new_rows = [row[:width] for row in new_rows]

        # Guard against silently wiping real business data: if the template
        # currently has fewer data rows than the CSV it's about to overwrite
        # (e.g. someone regenerated an empty/near-empty .xlsx template),
        # skip it instead of shrinking the file - this is exactly how
        # organizations.csv lost its 4 real rows in commit 87cb6cf.
        new_row_count = max(len(new_rows) - 1, 0)
        old_row_count = existing_data_row_count(csv_path)
        if not force and old_row_count > 0 and new_row_count < old_row_count:
            print(
                f"SKIP {excel_path.name}: would shrink {csv_path.name} from "
                f"{old_row_count} to {new_row_count} data rows - pass --force to overwrite anyway."
            )
            continue

        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerows(new_rows)

        print(f"{excel_path.name} -> {csv_path.relative_to(output_dir.parent)}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Directory containing .xlsx files",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Directory to write .csv files to",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite a CSV even if it would shrink the number of data rows",
    )
    args = parser.parse_args()

    convert(args.input.resolve(), args.output.resolve(), force=args.force)


if __name__ == "__main__":
    main()
